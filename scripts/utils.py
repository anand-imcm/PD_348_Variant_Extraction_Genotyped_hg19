#!/usr/bin/env python3

from openpyxl.styles import PatternFill, Font, Alignment

red_fill = PatternFill(start_color="F8C8DC", end_color="F8C8DC", fill_type="solid")
dark_red_font = Font(color="8B0000")

green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
dark_green_font = Font(color="006400")

def xlHighlight(worksheet):
    for row in worksheet.iter_rows(min_row=8, max_row=8, min_col=2, max_col=worksheet.max_column):
        for cell in row:
            try:
                if cell.value is not None and int(cell.value) > 1:
                    cell.fill = red_fill
                    cell.font = dark_red_font
            except (ValueError, TypeError) as e:
                pass

    for row in worksheet.iter_rows(min_row=9, max_row=12, min_col=2, max_col=worksheet.max_column):
        for cell in row:
            if cell.value is True:
                cell.fill = green_fill
                cell.font = dark_green_font

def xlFormat(worksheet):
    for col in worksheet.columns:
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = 40
    for cell in worksheet['A']:
        cell.alignment = Alignment(horizontal='left')